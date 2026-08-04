"""用户管理路由。

固定路径（/me /profile /options /password /mobile /email /template /export /import /status）
须在 /{user_id} 之前注册。
"""

import asyncio
import io

from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import StreamingResponse

from app.database import get_db
from app.dependencies import get_current_user, require_perm
from app.auth.schemas import SysUserDetails
from app.response import Result
from app.system.user.schemas import (
    EmailUpdateForm, ExcelResultVO, MobileUpdateForm, PasswordUpdateForm,
    PasswordVerifyForm, UserCreate, UserProfileForm, UserQuery,
    UserUpdate,
)
from app.system.user.service import UserService
from app.system.log.operation_log import operation_log
from app.system.log.constants import ActionTypeEnum, LogModuleEnum

router = APIRouter(prefix="/api/v1/users", tags=["用户管理"])


@router.get("", summary="用户分页列表", dependencies=[Depends(require_perm("sys:user:list"))])
async def get_user_page(
    pageNum: int = Query(default=1, ge=1),
    pageSize: int = Query(default=10, ge=1, le=100),
    keywords: str | None = Query(default=None),
    deptId: int | None = Query(default=None),
    status: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    user: SysUserDetails = Depends(get_current_user),
):
    query = UserQuery(pageNum=pageNum, pageSize=pageSize, keywords=keywords, deptId=deptId, status=status)
    return Result(data=await UserService(db).get_page(query, user))


@router.get("/me", summary="获取当前登录用户信息")
async def get_current_user_info(
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.auth.service import AuthService
    return Result(data=await AuthService(db).get_user_info(user.userId))


@router.get("/profile", summary="获取个人中心用户信息")
async def get_user_profile(
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return Result(data=await UserService(db).get_user_profile(user.userId))


@router.put("/profile", summary="个人中心修改用户信息")
@operation_log(module=LogModuleEnum.USER, action_type=ActionTypeEnum.UPDATE, title="修改个人资料")
async def update_user_profile(
    request: Request,
    form: UserProfileForm,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return Result(data=await UserService(db).update_user_profile(user.userId, form))


@router.put("/password", summary="当前用户修改密码")
@operation_log(module=LogModuleEnum.USER, action_type=ActionTypeEnum.CHANGE_PASSWORD, title="修改密码")
async def change_current_user_password(
    request: Request,
    form: PasswordUpdateForm,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).change_password(user.userId, form.oldPassword, form.newPassword)
    return Result(data=None)


@router.post("/mobile/code", summary="发送手机号验证码")
async def send_mobile_code(mobile: str, user: SysUserDetails = Depends(get_current_user)):
    # TODO: 接入真实短信服务
    return Result(data=None)


@router.put("/mobile", summary="绑定或更换手机号")
async def bind_or_change_mobile(
    form: MobileUpdateForm,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).bind_or_change_mobile(user.userId, form.mobile, form.smsCode)
    return Result(data=None)


@router.delete("/mobile", summary="解绑手机号")
async def unbind_mobile(
    form: PasswordVerifyForm,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).unbind_mobile(user.userId, form.password)
    return Result(data=None)


@router.post("/email/code", summary="发送邮箱验证码")
async def send_email_code(email: str, user: SysUserDetails = Depends(get_current_user)):
    # TODO: 接入真实邮件服务
    return Result(data=None)


@router.put("/email", summary="绑定或更换邮箱")
async def bind_or_change_email(
    form: EmailUpdateForm,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).bind_or_change_email(user.userId, form.email, form.smsCode)
    return Result(data=None)


@router.delete("/email", summary="解绑邮箱")
async def unbind_email(
    form: PasswordVerifyForm,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).unbind_email(user.userId, form.password)
    return Result(data=None)


@router.get("/options", summary="用户下拉选项")
async def get_user_options(db: AsyncSession = Depends(get_db)):
    return Result(data=await UserService(db).get_user_options())


@router.get("/template", summary="下载用户导入模板", dependencies=[Depends(require_perm("sys:user:import"))])
async def download_template():
    loop = asyncio.get_running_loop()

    def _build():
        wb = Workbook()
        ws = wb.active
        ws.title = "用户导入模板"
        ws.append(["用户名", "昵称", "手机号", "邮箱", "性别", "状态"])
        ws.append(["zhangsan", "张三", "13800138000", "zhangsan@example.com", "0", "启用"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = await loop.run_in_executor(None, _build)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_template.xlsx"},
    )


@router.get("/export", summary="导出用户", dependencies=[Depends(require_perm("sys:user:export"))])
@operation_log(module=LogModuleEnum.USER, action_type=ActionTypeEnum.EXPORT, title="导出用户")
async def export_users(
    request: Request,
    keywords: str | None = Query(default=None),
    deptId: int | None = Query(default=None),
    status: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    user: SysUserDetails = Depends(get_current_user),
):
    data = await UserService(db).export_users(UserQuery(keywords=keywords, deptId=deptId, status=status), user)
    loop = asyncio.get_running_loop()

    def _build():
        wb = Workbook()
        ws = wb.active
        ws.title = "用户列表"
        if data:
            ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = await loop.run_in_executor(None, _build)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.post("/import", summary="导入用户", dependencies=[Depends(require_perm("sys:user:import"))])
@operation_log(module=LogModuleEnum.USER, action_type=ActionTypeEnum.IMPORT, title="导入用户")
async def import_users(
    request: Request,
    file: UploadFile = File(...),
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    loop = asyncio.get_running_loop()
    content = await file.read()

    def _parse():
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_data = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows_data:
            return []
        headers = [str(c) if c is not None else "" for c in rows_data[0]]
        data = []
        for row in rows_data[1:]:
            if any(v is not None for v in row):
                data.append(dict(zip(headers, row, strict=False)))
        return data

    data = await loop.run_in_executor(None, _parse)
    result = await UserService(db).import_users(data)
    return Result(data=ExcelResultVO(**result))


# ── 路径参数端点（最后注册）──

@router.post("", summary="创建用户", dependencies=[Depends(require_perm("sys:user:create"))])
async def create_user(
    form: UserCreate,
    db: AsyncSession = Depends(get_db),
    user: SysUserDetails = Depends(get_current_user),
):
    return Result(data=await UserService(db).create(form, user.userId))


@router.get("/{user_id}", summary="用户详情", dependencies=[Depends(require_perm("sys:user:detail"))])
async def get_user(user_id: int, db: AsyncSession = Depends(get_db)):
    return Result(data=await UserService(db).get_by_id(user_id))


@router.get("/{user_id}/form", summary="用户表单数据", dependencies=[Depends(require_perm("sys:user:update"))])
async def get_user_form(user_id: int, db: AsyncSession = Depends(get_db)):
    return Result(data=await UserService(db).get_user_form(user_id))


@router.put("/{user_id}", summary="更新用户", dependencies=[Depends(require_perm("sys:user:update"))])
async def update_user(
    user_id: int,
    form: UserUpdate,
    db: AsyncSession = Depends(get_db),
    user: SysUserDetails = Depends(get_current_user),
):
    form.id = user_id
    return Result(data=await UserService(db).update(form, user.userId))


@router.delete("/{ids}", summary="删除用户", dependencies=[Depends(require_perm("sys:user:delete"))])
async def delete_users(ids: str, db: AsyncSession = Depends(get_db)):
    count = await UserService(db).delete(ids)
    return Result(data=count, msg=f"成功删除 {count} 条记录")


@router.patch("/{user_id}/status", summary="修改用户状态", dependencies=[Depends(require_perm("sys:user:update"))])
@operation_log(module=LogModuleEnum.USER, action_type=ActionTypeEnum.UPDATE, title="修改用户状态")
async def update_user_status(
    request: Request,
    user_id: int,
    status: int,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).update_status(user_id, status)
    return Result(data=None)


@router.put("/{user_id}/password/reset", summary="重置指定用户密码", dependencies=[Depends(require_perm("sys:user:reset-password"))])
@operation_log(module=LogModuleEnum.USER, action_type=ActionTypeEnum.RESET_PASSWORD, title="重置用户密码")
async def reset_user_password(
    request: Request,
    user_id: int,
    password: str,
    user: SysUserDetails = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await UserService(db).reset_password(user_id, password)
    return Result(data=None)
